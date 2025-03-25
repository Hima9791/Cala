import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
import time
from tqdm import tqdm
from io import BytesIO
import gc  # Added for memory management

# ========================
# --- QA Check Functions ---
# ========================

def create_key(df):
    df['Key'] = df['ChemicalID'].astype(str) + '_' + df['PartNumber'].astype(str)
    return df

def validate_rows_count(df):
    actual_counts = df['Key'].value_counts().reset_index()
    actual_counts.columns = ['Key', 'ActualRowsCount']
    df = df.merge(actual_counts, on='Key')
    df['RowsCountGap'] = df['RowsCount '] - df['ActualRowsCount']
    df['Automated QA Comment'] += df['RowsCountGap'].apply(lambda x: ' | Rows count mismatch' if x != 0 else '')
    return df

def check_fmd_revision_flag(df):
    df['Automated QA Comment'] += df['FMDRevFlag'].apply(lambda x: ' | FMDRevFlag is Not Latest' if x == 'Not Latest' else '')
    return df

def check_homogeneous_material_mass_variation(df):
    df['HomogeneousMaterialName'] = df['HomogeneousMaterialName'].str.lower()
    grouped = df.groupby(['Key', 'HomogeneousMaterialName'])
    for (key, material), group in grouped:
        if group['HomogeneousMaterialMass '].nunique() > 1:
            idx = (df['Key'] == key) & (df['HomogeneousMaterialName'] == material)
            df.loc[idx, 'Automated QA Comment'] += ' | Multiple masses for the same homogeneous material'
    return df

def check_homogeneous_material_mass(df):
    df['CalculatedMass'] = df.groupby(['Key', 'HomogeneousMaterialName'])['Mass '].transform('sum')
    df['Homogeneous Mass Gap'] = df['CalculatedMass'] - df['HomogeneousMaterialMass ']
    df['Automated QA Comment'] += df['Homogeneous Mass Gap'].apply(lambda x: ' | Fail: Mass mismatch' if abs(x) >= 1 else '')
    return df

def check_substance_homogeneous_material_percentage(df):
    sums = df.groupby(['Key', 'HomogeneousMaterialName'])['SubstanceHomogeneousMaterialPercentage '].transform('sum')
    df['Automated QA Comment'] += sums.apply(lambda x: '' if 99.9 <= x <= 100.1 else ' | Fail: homogeneousPercentage sum != 100')
    return df

def check_substance_homogeneous_material_ppm(df):
    sums = df.groupby(['Key', 'HomogeneousMaterialName'])['SubstanceHomogeneousMaterialPercentagePPM '].transform('sum')
    df['Automated QA Comment'] += sums.apply(lambda x: '' if 999000 <= x <= 1001000 else ' | Fail: homogeneousPPM sum != 1000000')
    return df

def check_substance_component_level_percentage(df):
    sums = df.groupby('Key')['SubstanceComponentLevelPercentage '].transform('sum')
    df['Automated QA Comment'] += sums.apply(lambda x: '' if 99.0 <= x <= 101.0 else ' | Fail: Component level percentage sum != 100')
    return df

def check_substance_component_level_ppm(df):
    sums = df.groupby('Key')['SubstanceComponentLevelPPM '].transform('sum')
    df['Automated QA Comment'] += sums.apply(lambda x: '' if 990000 <= x <= 1010000 else ' | Fail: Component level PPM sum != 1000000')
    return df

def calculate_gap_and_comment(df):
    gap_percentage = abs(df['TotalComponentMassProfile '] - df['TotalComponentMassSummation ']) / df['TotalComponentMassProfile '] * 100
    df['Automated QA Comment'] += gap_percentage.apply(lambda x: ' | Total VS Summation Gap is more than 50%' if x >= 50 else '')
    return df

def check_total_component_mass_summation(df):
    for key in df['Key'].unique():
        group = df[df['Key'] == key]
        if round(group['Mass '].sum(), 4) != group['TotalComponentMassSummation '].iloc[0]:
            df.loc[df['Key'] == key, 'Automated QA Comment'] += ' | Software issue'
    return df

# ========================
# --- Chunk Processing ---
# ========================

def process_chunk(chunk):
    checks = [
        check_fmd_revision_flag,
        check_homogeneous_material_mass_variation,
        validate_rows_count,
        check_homogeneous_material_mass,
        check_substance_homogeneous_material_percentage,
        check_substance_homogeneous_material_ppm,
        check_substance_component_level_percentage,
        check_substance_component_level_ppm,
        calculate_gap_and_comment,
        check_total_component_mass_summation
    ]

    if 'Automated QA Comment' not in chunk.columns:
        chunk['Automated QA Comment'] = ''

    for check in checks:
        chunk = check(chunk)

    # Clean temporary columns to save RAM
    temp_cols = ['CalculatedMass', 'RowsCountGap', 'Homogeneous Mass Gap']
    chunk.drop(columns=temp_cols, inplace=True, errors='ignore')

    return chunk

def run_all_checks(file_data):
    start_time = time.time()
    workbook = openpyxl.load_workbook(file_data, read_only=True)
    sheet = workbook.active
    df_all = pd.DataFrame(sheet.values)
    df_all.columns = df_all.iloc[0]
    df_all = df_all[1:]

    df_all = create_key(df_all)
    unique_keys = df_all['Key'].unique()

    chunk_size = 20  # Adjust for RAM efficiency
    final_columns = df_all.columns.tolist() + ['Automated QA Comment']

    output_wb = openpyxl.Workbook(write_only=True)
    output_sheet = output_wb.create_sheet()

    # Write header
    header = [openpyxl.cell.WriteOnlyCell(output_sheet, value=col) for col in final_columns]
    header[-1].font = Font(color="FF0000")
    output_sheet.append(header)

    for i in tqdm(range(0, len(unique_keys), chunk_size)):
        chunk_keys = unique_keys[i:i+chunk_size]
        chunk = df_all[df_all['Key'].isin(chunk_keys)].copy()
        chunk = process_chunk(chunk)
        # Convert each row to a list to satisfy openpyxl's requirements
        for row in chunk[final_columns].values:
            output_sheet.append(row.tolist())
        del chunk
        gc.collect()

    buffer = BytesIO()
    output_wb.save(buffer)
    buffer.seek(0)

    mins, secs = divmod(time.time() - start_time, 60)
    st.write(f"Process Complete. Elapsed Time: {int(mins):02}:{int(secs):02}")

    return buffer

# ========================
# --- Streamlit App UI ---
# ========================

def main():
    hide_elements = """
    <style>
    #MainMenu {visibility: hidden;}
    footer a[href*="github.com"] {display: none !important;}
    div[class^="_profilePreview"] {display: none !important;}
    </style>
    """
    st.markdown(hide_elements, unsafe_allow_html=True)
    st.title("QA Checker for Chemical Smart Checkers")
    answer = st.radio("Is Ibrahem a good person?", ("Yes", "No"))
    if answer == "No":
        st.warning("ok, check it manually")
    else:
        uploaded_file = st.file_uploader("Upload an Excel file", type=["xlsx", "xls"])
        if uploaded_file is not None:
            try:
                result_buffer = run_all_checks(uploaded_file)
                st.success("File processed successfully!")
                st.download_button("Download Processed File", result_buffer, "output_checked.xlsx")
            except Exception as e:
                st.error(f"Error: {e}")

if __name__ == '__main__':
    main()
